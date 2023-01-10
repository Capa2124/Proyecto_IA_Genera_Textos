from PyPDF2 import PdfReader
from datetime import datetime
import re
import os
from os import system
from palabra import *
import xlsxwriter as xl
import openpyxl

def lector(inicio:int, fin:int):
    nombreLibro = "Libro_Completo"
    if(not os.path.isfile(nombreLibro + str(inicio) + str(fin) + ".txt")):
        reader = PdfReader(nombreLibro + ".pdf")
        for i in range(inicio,fin):
            page = reader.pages[i]
            pagina = page.extract_text()
            if i == inicio:
                libro = pagina
            else:
                libro = libro + " " + pagina
        file = open(nombreLibro + str(inicio) + str(fin) + ".txt", "w")
        libro = limpieza(libro)
        file.write(libro)
    else:
        file = open(nombreLibro + str(inicio) + str(fin) + ".txt", "r")
        libro = file.read()
    file.close()
    #escribe_bitacora("Se leyo el PDF y se extrajo el texto desde la página " +  str(inicio) + " hasta " + str(fin) + "\n")
    return libro

def limpieza_saltoLinea(text:str):
    text = re.sub("\n","",text)
    return text

def limpieza(text:str):
    carateresEspeciales = "—-_'.!¡¿?…:;«»“”*|°""/\~()[]{}><•,"
    text = text.lower()#Coloca todo el texto en minusculas
    #text = repar_palabras(text) #Repara algunas palabras 
    text = ''.join(x for x in text if x not in carateresEspeciales)#Elimina todos los carateres especiales
    text = re.sub("\n"," ",text)#Elimina todos los saltos de linea
    text = re.sub("\t"," ",text)#Elimina todos los Tabs
    text = re.sub("  +"," ",text)#Elimina todos los 2bles espacios o mas
    text = repar_palabras(text) #Repara algunas palabras 
    #escribe_bitacora("Se realizo la limpieza del texto.\n")
    return text

def limpieza2(text:str):
    carateresEspeciales = "[]()'"
    text = ''.join(x for x in text if x not in carateresEspeciales)#Elimina todos los carateres especiales
    text.split(", ")
    return text


def repar_palabras(libro_aReparar:str):
    libro_aReparar = libro_aReparar.replace(' v ', ' v')
    libro_aReparar = libro_aReparar.replace('llév atelo', 'llévatelo')
    libro_aReparar = libro_aReparar.replace('vuelv a', 'vuelva')
    libro_aReparar = libro_aReparar.replace('educativ a', 'educativa')
    libro_aReparar = libro_aReparar.replace('v oldemort', 'voldemort')
    libro_aReparar = libro_aReparar.replace(' y y a ', ' y a ')
    libro_aReparar = libro_aReparar.replace(' , ', ', ')
    for i in range(1, 783):
        libro_aReparar = re.sub("wwwlectulandiacom página " + str(i), "", libro_aReparar)
        libro_aReparar = re.sub("capítulo " + str(i),"", libro_aReparar)
    return libro_aReparar


def mejora_oracion(oracion):
    oracion = oracion.capitalize()
    #if oracion[:-1] == " ":
    oracion = oracion[:-1]
    oracion = oracion + "."
    escribe_bitacora("\nSe mejoro la oracion.\n")
    return oracion

def eliminaPreposiciones(patron:str):
    listaPrepociones = ["a", "ante", "bajo", "cabe", "con", "contra", "de", "desde", "durante", "en", "entre",
    "hacia", "hasta", "mediante", "para", "por", "según", "sin", "so", "sobre", "tras", "versus", "vía"]
    patron = patron.split()
    for i in patron:
        #Elimina las prepociones
        for j in listaPrepociones:
            rep1 = patron.count(j)
            for k in range(1, rep1+1):
                if i == j:
                    patron.remove(j)
    return juntarSplit(patron)

def eliminaArticulos(patron:str):
    listaArticulos = ["las", "la", "los", "lo", "el", "un", "una", "unos", "unas"]
        #Elimina los Articulos
    patron = patron.split()
    for i in patron:
        for j in listaArticulos:
            rep2 = patron.count(j)
            for k in range(1, rep2+1):
                if i == j:
                    patron.remove(j)
    return juntarSplit(patron)

def juntarSplit(patron):
    patronJunto = ""
    for i in patron:
        if "" == patronJunto:
            patronJunto = i
        else:   
            patronJunto = patronJunto + " " + i
    return patronJunto

def referenciaSigElemento(ListaPatron:list, dicccionario, referenciaPatron:str, similarPatrones:int, probabilidad:str): #Lista_Patron = regresa busqueda_patron, referenciaPatron = referenciaUsuario, #similarPatrones = .5
    case = 1
    bandera = True
    while bandera:
        if case == 1:
            if referenciaPatron in ListaPatron:
                bandera = False 
                listaReferencia  = referenciaPatron.split()
                if len(dicccionario[listaReferencia[-1]].tupla_sig_patron) == 0:
                    return dicccionario[listaReferencia[-1]].get_sig_palabra(probabilidad)
                else:
                    return dicccionario[listaReferencia[-1]].get_sig_patron(probabilidad)
            elif len(referenciaPatron.split()) == 1:
                bandera = False 
                listaReferencia  = referenciaPatron.split()
                return dicccionario[listaReferencia[-1]].get_sig_palabra(probabilidad)
            else:
                case = 2
        elif case == 2:
            auxreferenciaPatron = eliminaArticulos(referenciaPatron)
            auxreferenciaPatron = eliminaPreposiciones(auxreferenciaPatron)
            numPalabrasUnicas = len(auxreferenciaPatron.split())
            for i in ListaPatron: 
                contador = 0
                iaux = eliminaArticulos(i)
                iaux = eliminaPreposiciones(iaux)
                for j in iaux.split():#Revisar cada palabra unica
                    if j in auxreferenciaPatron:
                        contador+=1
                numero = numPalabrasUnicas*(similarPatrones*.01)#Obtenemos el porcentaje 
                redondeo = round(numero)   
                if contador >= redondeo and contador != 0:
                    case = 1
                    referenciaPatron = i
                    break
                else:
                    case = 3
        elif case == 3:
            referenciaPatron = referenciaPatron.split()
            case = 1
            if len(referenciaPatron) > 1:
                referenciaPatron.pop()
                referenciaPatron = juntarSplit(referenciaPatron)

def escribe_bitacora(registro):
    now = datetime.now()
    nombre_archivo = str(now.day) + '-' + str(now.month) + '-' + str(now.year) + '_' +  str(now.hour) + 'horas_' + str(now.minute) + 'minutos_'
    file = open(nombre_archivo + '.txt', "a")
    file.write(registro)
    file.close()

def separa_texto(texto):
    return texto.split(" ")


def agrega_sig_elemento(lista_palabras, lista_inicio_patron, diccionario_patrones):
    # función capaz de obtener las dos primeras palabras y 
    # analizar si forman parte de un patrón ya existente y cuántas veces continúa al patrón anterior  
    # En caso de que no entonces la palabra será incluida en la lista sig palabra 
    diccionario = {}
    bandera = False
    #lista_palabras=["Harry", "ron", "hermione", "estaban", "comiendo"]
    # ron hermione ...
    # lista_palabras= Harry ron hermione y harry ron hermione y harry
    # diccionario_capa = {"harry potter ron":3, "ron hermione y voldemort":4, "sin ti":4, "contigo no":4, "lord voldemort":6}
    # lista_ultima_palabra = [[2, 12, 223],[987, 543, 123, 443, ],[ 4432, 223, 554,432], [324, 4324, 223,44], [4432, 443, 5454, 6565, 343,555]]
    keys_diccionario_patrones = list(diccionario_patrones.keys())
    #[palabra, {sig_patron : aparicion}; {sig_palabra: aparicion}]
    for i in range(len(lista_palabras)-1): #LEEMOS TODAS LAS PALABRAS DEL TEXTO
        
        if lista_palabras[i] in diccionario.keys():
            diccionario[lista_palabras[i]].num_ocurrencia += 1
        else:
            diccionario[lista_palabras[i]] = Palabra(lista_palabras[i])
        for j in range(0, len(lista_inicio_patron)): #ITERAMOS DENTRO DE LA LISTA PRINCIPAL
            if i+1 in lista_inicio_patron[j]:
                patron_actual = keys_diccionario_patrones[j]
                bandera = True
                break
        if bandera:
            diccionario[lista_palabras[i]].agrega_sig_patron(patron_actual)
            diccionario[lista_palabras[i]].agrega_sig_palabra(lista_palabras[i+1])
            bandera = False
        else:
            diccionario[lista_palabras[i]].agrega_sig_palabra(lista_palabras[i+1]) 
    for i in diccionario.keys():
        diccionario[i].cierra_tabla()
    escribe_bitacora("Se crea un diccionario con todos los elementos del texto y sus siguientes elementos.\n")
    return diccionario
        
def crea_oracion(diccionario, tamanio, probabilidad, referencia = "harry"):
    textoFinal = ""
    for i in range(int(tamanio)):
        escribe_bitacora("Referencia: " +  referencia + "  sig palabras: " + str(diccionario[referencia].tupla_sig_palabra) + "\n\n")
        textoFinal += referencia + " "
        aux = diccionario[referencia]
        referencia = aux.get_sig_palabra(probabilidad)
    return textoFinal

def crea_Narrativa(ListaPatron:list, diccionario, numeroIteraciones:int, similarPatrones:int, referencia:str, probabilidad:str):
    textoFinal = ""
    for i in range(1, numeroIteraciones + 1):
        if i == 1:
            textoFinal = referencia
        else:
            sigReferencia = referenciaSigElemento(ListaPatron, diccionario, referencia, similarPatrones, probabilidad)
            textoFinal = textoFinal + " " + sigReferencia
            referencia = sigReferencia
    return textoFinal

def guardarPatrones(ListaPatrones:list):
    file = open("patrones.txt", "w")
    for i in ListaPatrones:
        file.write(i + "\n")
    file.close()

def leePatrones():
    file = open("patrones.txt", "r")
    patron = file.readlines()
    listaPatron = []
    for i in patron:
        listaPatron.append(limpieza_saltoLinea(i))
    file.close()
    return listaPatron

def crea_excel(diccionario):
    archivo = xl.Workbook('Base_datos.xlsx')
    hoja=archivo.add_worksheet()
    i = 0
    for llave, valor in diccionario.items():
        hoja.write(i,0, llave)
        hoja.write(i,1, valor.num_ocurrencia+1)
        hoja.write(i,2, str(valor.tupla_sig_palabra))
        hoja.write(i,3, str(valor.tupla_sig_patron))
        i+=1
    archivo.close()

def recursiva_busca_patron(base, texto, busqueda, dict):
    patron_n = es_patron(base, texto)
    if patron_n[0]:
        if busqueda+1 < len(texto):
            base.append(texto[busqueda])
            ayuda = recursiva_busca_patron(base, texto, busqueda+1, dict)
            if ayuda[1]:
                meter = " ".join(base[:-1])
                if meter not in dict:
                    dict[meter] = patron_n[1]
                return busqueda-1, False, dict, patron_n[2]
            return ayuda
        else:
            return busqueda-1, False, dict
    else: #YA NO ES PATRON O NO FUE DESDE EL INICIO
        return busqueda-1, True, dict
        
def es_patron(list1, list2):
    indice1 = 0
    patron = 0
    indices_p = []
    for i in range(len(list2)):
        if list1[indice1] == list2[i]:
            aux = i
            for j in range(len(list1)):
                if list1[j] == list2[aux]:
                    indice1 += 1
                    if aux+1 <len(list2):
                        aux += 1
                else:
                    break
            if indice1 == len(list1):
                indices_p.append(aux-len(list1))
                indice1 = 0
                patron += 1
            else:
                indice1 = 0
        else:
            continue
    if patron >= 3:
        return True, patron, indices_p
    else:
        return False, patron, indices_p

def busqueda_patrones(texto):
    prueba = [0,False,{}]
    i=0
    li_li = []
    while i < len(texto):
        if i+2<len(texto):
            prueba = recursiva_busca_patron([texto[i], texto[i+1]], texto, i+2, prueba[2])
            if len(prueba)==4 and prueba[3] not in li_li:
                li_li.append(prueba[3])

            i=prueba[0]
        else:
            i+=1
    return(prueba[2], li_li)

def lee_excel():
    dict = {}
    lista_final = []
    lista_finalP = []
    li_t = ()
    li_p = ()
    i=1
    dataframe = openpyxl.load_workbook("Base_datos.xlsx")
    # Define variable to read sheet
    dataframe1 = dataframe.active
    # Iterate the loop to read the cell values
    for row in range(0, int(dataframe1.max_row)):
        for columna in dataframe1.iter_cols(1, dataframe1.max_column):
            if i == 1:
                dict[columna[row].value]=Palabra(columna[row].value)
                llave = columna[row].value
                i+=1
            elif i == 2:
                dict[llave].num_ocurrencia=columna[row].value
                i+=1
            elif i==3:
                dict[llave].diccionario_sig_palabra = columna[row].value
                listaP=limpieza2(columna[row].value).split(", ")
                for z in range(1,len(listaP),2):
                    li_p = (listaP[z-1], float(listaP[z]))
                    lista_finalP.append(li_p)
                dict[llave].tupla_sig_palabra = lista_finalP
                lista_finalP = []
                i+=1
            else:
                dict[llave].diccionario_sig_patron = columna[row].value
                #dict[llave].diccionario_sig_patron = list((columna[row].value).split(" "))
                listaL=limpieza2(columna[row].value).split(", ")
                for x in range(1,len(listaL),2):
                    li_t = (listaL[x-1], float(listaL[x]))
                    lista_final.append(li_t)
                    #print(lista_final)
                #print("Lista final: ", lista_final)
                dict[llave].tupla_sig_patron = lista_final
                lista_final = []
        i=1
    return dict
